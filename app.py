"""
Credit Management System for Skanda Enterprises
Main Flask application file
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
import os
import io
import uuid
import json
from functools import wraps
from sqlalchemy import inspect
from sqlalchemy import text
import traceback as tb
from PIL import Image
try:
    import easyocr
    EASYOCR_AVAILABLE = True
    # Cache the reader to avoid re-initializing (first call downloads models)
    _easyocr_reader = None
    _easyocr_initializing = False
except ImportError:
    EASYOCR_AVAILABLE = False
    _easyocr_reader = None
    _easyocr_initializing = False
from ocr_utils import parse_bill_data, compare_bill_data, parse_bill_number, parse_amount, parse_date, parse_vendor_name
from vendor_ocr_utils import match_vendor_from_ocr, extract_vendor_name_from_text
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'skanda-enterprises-credit-management-2024')

# Database configuration for Vercel compatibility
database_url = os.environ.get('DATABASE_URL', 'sqlite:///credit_management.db')
# Handle Vercel Postgres URL format (postgres:// -> postgresql://)
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
# For SQLite on Vercel, use /tmp directory (only writable location)
if database_url.startswith('sqlite:///') and os.environ.get('VERCEL'):
    db_name = database_url.replace('sqlite:///', '')
    database_url = f'sqlite:////tmp/{db_name}'

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# For Vercel, use /tmp for uploads (only writable location in serverless)
if os.environ.get('VERCEL'):
    app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', '/tmp/uploads')
else:
    app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', 'static/uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

db = SQLAlchemy(app)

# OCR Functions - EasyOCR Only
def extract_text_from_image(image_path):
    """Extract text from image using EasyOCR"""
    if not EASYOCR_AVAILABLE:
        return "", 0, "EasyOCR is not installed. Please install it with: pip install easyocr"
    
    try:
        # Check if file exists
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image file not found: {image_path}")
        
        # Handle PDF files (convert to image first)
        file_ext = os.path.splitext(image_path)[1].lower()
        if file_ext == '.pdf':
            return "", 0, "PDF files are not directly supported. Please convert PDF to image (PNG/JPEG) first."
        
        # Verify image is valid
        try:
            image = Image.open(image_path)
            image.verify()
        except Exception as img_error:
            return "", 0, f"Invalid image file: {str(img_error)}"
        
        # Use cached reader or initialize (first call downloads models)
        global _easyocr_reader, _easyocr_initializing
        
        if _easyocr_reader is None and not _easyocr_initializing:
            print("Initializing EasyOCR reader (first time may take a few minutes to download models)...")
            _easyocr_initializing = True
            try:
                _easyocr_reader = easyocr.Reader(['en'], gpu=False, verbose=False)
                print("EasyOCR reader initialized successfully!")
                _easyocr_initializing = False
            except Exception as init_error:
                _easyocr_initializing = False
                error_msg = f"EasyOCR initialization failed: {str(init_error)}"
                print(f"OCR Error: {error_msg}")
                return "", 0, error_msg
        
        if _easyocr_reader is None:
            return "", 0, "EasyOCR reader initialization failed. Please restart the application."
        
        # Extract text using EasyOCR
        print(f"Processing image with EasyOCR: {image_path}")
        result = _easyocr_reader.readtext(image_path)
        
        # Combine all detected text
        text_lines = [detection[1] for detection in result]
        text = '\n'.join(text_lines)
        
        # Calculate average confidence
        confidences = [detection[2] for detection in result]
        avg_confidence = sum(confidences) * 100 / len(confidences) if confidences else 0
        
        print(f"EasyOCR extracted {len(text_lines)} text lines, confidence: {avg_confidence:.1f}%")
        
        if not text.strip():
            return "", 0, "OCR extracted no text. Please ensure the image is clear and contains readable text."
        
        return text.strip(), avg_confidence, None
        
    except FileNotFoundError as e:
        error_msg = f"Image file not found: {str(e)}"
        print(f"OCR Error: {error_msg}")
        return "", 0, error_msg
    except Exception as e:
        error_msg = f"OCR Error: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        return "", 0, error_msg

def allowed_file(filename):
    """Check if file extension is allowed"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff', 'pdf'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Role Constants
ROLES = {
    'admin': 'Admin',
    'delivery_man': 'Delivery Man',
    'salesman': 'Salesman',
    'computer_organiser': 'Computer Organiser'
}

# Permission Definitions
PERMISSIONS = {
    "admin": ["*"],  # Admin has all permissions
    "salesman": ["bills.create", "bills.update", "credits.create", "vendors.view"],
    "computer_organiser": ["vendors.update", "vendors.view", "ocr.verify", "reports.export", "bills.verify"],
    "delivery_man": ["deliveries.update_status", "deliveries.view"]
}

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='delivery_man')
    full_name = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    invoices = db.relationship('Invoice', backref='creator', lazy=True)
    credits = db.relationship('Credit', backref='creator', lazy=True)
    bills = db.relationship('Bill', foreign_keys='Bill.created_by', backref='creator', lazy=True)
    delivery_orders = db.relationship('DeliveryOrder', backref='delivery_man', lazy=True)
    activities = db.relationship('ActivityLog', backref='user', lazy=True)
    audit_logs = db.relationship('AuditLog', backref='user', lazy=True)
    
    def get_permissions(self):
        """Get list of permission codes for this user's role"""
        if self.role == 'admin':
            return ['*']  # Admin has all permissions
        role = Role.query.filter_by(name=self.role).first()
        if role:
            return [perm.code for perm in role.permissions]
        return []
    
    def has_permission(self, permission_code):
        """Check if user has a specific permission"""
        user_perms = self.get_permissions()
        if '*' in user_perms:
            return True
        return permission_code in user_perms
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def is_admin(self):
        return self.role == 'admin'
    
    def is_delivery_man(self):
        return self.role == 'delivery_man'
    
    def is_salesman(self):
        return self.role == 'salesman'
    
    def is_computer_organiser(self):
        return self.role == 'computer_organiser'

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    company_name = db.Column(db.String(100))
    gst_number = db.Column(db.String(20))
    credit_limit = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship
    invoices = db.relationship('Invoice', backref='customer', lazy=True)

class Salesman(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship
    invoices = db.relationship('Invoice', backref='salesman', lazy=True)

class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(50), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey('salesman.id'), nullable=False)
    delivery_date = db.Column(db.Date, nullable=False)
    bill_amount = db.Column(db.Float, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Image fields for OCR
    image_filename = db.Column(db.String(255), nullable=True)
    extracted_text = db.Column(db.Text, nullable=True)
    ocr_confidence = db.Column(db.Float, nullable=True)
    
    # Relationships
    credits = db.relationship('Credit', backref='invoice', lazy=True, cascade='all, delete-orphan')
    deliveries = db.relationship('DeliveryOrder', backref='invoice', lazy=True)
    
    @property
    def total_credits(self):
        return sum(credit.amount for credit in self.credits)
    
    @property
    def outstanding_balance(self):
        return self.bill_amount - self.total_credits
    
    @property
    def is_paid(self):
        return self.outstanding_balance <= 0

class Vendor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    vendor_name = db.Column(db.String(100))  # Alias for name (for compatibility)
    vendor_type = db.Column(db.String(50))  # Wholesale, Retail, Supplier, Manufacturer
    contact_number = db.Column(db.String(20))  # Alias for contact
    contact = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    gst_number = db.Column(db.String(20))
    category = db.Column(db.String(50))  # e.g., Supplier, Manufacturer (alias for vendor_type)
    
    # Credit tracking fields
    total_credit = db.Column(db.Numeric(10, 2), default=0.0)
    outstanding_credit = db.Column(db.Numeric(10, 2), default=0.0)
    cleared_credit = db.Column(db.Numeric(10, 2), default=0.0)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    bills = db.relationship('Bill', backref='vendor', lazy=True)
    
    def update_credit_totals(self):
        """Recalculate credit totals from linked credit transactions"""
        credits = self.credit_transactions or []
        self.total_credit = sum(float(c.credit_amount) for c in credits)
        self.outstanding_credit = sum(float(c.credit_amount) for c in credits if c.status in ['Pending', 'Overdue'])
        self.cleared_credit = sum(float(c.credit_amount) for c in credits if c.status == 'Cleared')
    

class Bill(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bill_number = db.Column(db.String(50), unique=True, nullable=False)
    bill_type = db.Column(db.String(20), nullable=False)  # 'handbill' or 'normal'
    vendor_id = db.Column(db.Integer, db.ForeignKey('vendor.id'), nullable=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey('salesman.id'), nullable=True)
    amount = db.Column(db.Float, nullable=False)
    bill_date = db.Column(db.Date, nullable=False)
    
    # Payment information
    payment_method = db.Column(db.String(50), default='Cash')  # Cash, Credit, UPI, Bank
    
    # OCR fields
    image_filename = db.Column(db.String(255))
    extracted_text = db.Column(db.Text)  # Full OCR extracted text
    ocr_text = db.Column(db.Text)  # Alias for extracted_text, kept for compatibility
    ocr_confidence = db.Column(db.Float)
    is_verified = db.Column(db.Boolean, default=False)
    verified_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    verified_at = db.Column(db.DateTime)
    
    # Status and verification
    status = db.Column(db.String(20), default='pending')  # pending, verified, rejected, discrepancy_found
    verification_status = db.Column(db.String(20), default='unverified')  # unverified, verified, discrepancy_found
    
    # Parsed OCR data (for comparison)
    ocr_bill_number = db.Column(db.String(50))
    ocr_amount = db.Column(db.Float)
    ocr_date = db.Column(db.Date)
    ocr_vendor_name = db.Column(db.String(200))
    
    notes = db.Column(db.Text)
    
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships (vendor relationship already exists via backref in Vendor model)
    salesman_rel = db.relationship('Salesman', backref='bills')
    credit_transactions = db.relationship('CreditTransaction', backref='bill_obj', foreign_keys='CreditTransaction.bill_id')

class DeliveryOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_number = db.Column(db.String(50), unique=True, nullable=False)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=True)
    delivery_man_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Delivery details
    delivery_date = db.Column(db.Date, nullable=False)
    expected_time = db.Column(db.Time)
    actual_delivery_time = db.Column(db.DateTime)
    
    # Status tracking
    status = db.Column(db.String(20), default='pending')  # pending, in_transit, delivered, cancelled
    delivery_address = db.Column(db.Text, nullable=False)
    customer_contact = db.Column(db.String(20))
    
    # OCR confirmation
    receipt_image = db.Column(db.String(255))
    receipt_ocr_text = db.Column(db.Text)
    is_confirmed = db.Column(db.Boolean, default=False)
    
    # Notes and remarks
    delivery_remarks = db.Column(db.Text)
    customer_signature = db.Column(db.Text)  # Store as base64 or file path
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Role(db.Model):
    """User Roles"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)  # admin, salesman, etc.
    display_name = db.Column(db.String(100))  # Admin, Salesman, etc.
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    permissions = db.relationship('Permission', secondary='role_permission', backref='roles', lazy='dynamic')

class Permission(db.Model):
    """Permissions"""
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(100), unique=True, nullable=False)  # bills.create, vendors.update, etc.
    description = db.Column(db.String(255))
    module = db.Column(db.String(50))  # bills, vendors, credits, etc.
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Association table for Role-Permission many-to-many
role_permission = db.Table('role_permission',
    db.Column('role_id', db.Integer, db.ForeignKey('role.id'), primary_key=True),
    db.Column('permission_id', db.Integer, db.ForeignKey('permission.id'), primary_key=True)
)

class ActivityLog(db.Model):
    """Legacy activity log - kept for backward compatibility"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action_type = db.Column(db.String(50), nullable=False)  # create, update, delete, approve, reject
    target_entity = db.Column(db.String(50), nullable=False)  # invoice, credit, delivery, vendor
    target_id = db.Column(db.Integer, nullable=False)
    old_value = db.Column(db.Text)
    new_value = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class AuditLog(db.Model):
    """Enhanced audit log with JSON metadata and trace_id"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # User's role at time of action
    action_type = db.Column(db.String(50), nullable=False)  # CREATE, UPDATE, DELETE, VERIFY, EXPORT
    target_type = db.Column(db.String(50), nullable=False)  # Vendor, Bill, Credit, OCR, Delivery
    target_id = db.Column(db.String(100), nullable=False)  # ID or unique key
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(255))
    meta = db.Column(db.Text)  # JSON string for metadata
    success = db.Column(db.Boolean, default=True)
    trace_id = db.Column(db.String(100), unique=True)  # Unique request trace
    
    def get_meta(self):
        """Parse meta JSON string"""
        if self.meta:
            try:
                return json.loads(self.meta)
            except:
                return {}
        return {}
    
    def set_meta(self, data):
        """Set meta as JSON string"""
        self.meta = json.dumps(data) if data else None

class Credit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_date = db.Column(db.Date, nullable=False)
    payment_method = db.Column(db.String(50), default='Cash')
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class OCRAuditLog(db.Model):
    """OCR Audit Log - Tracks re-verification and discrepancies"""
    id = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('bill.id'), nullable=False)
    verification_type = db.Column(db.String(50), nullable=False)  # initial, re_verification
    discrepancy_found = db.Column(db.Boolean, default=False)
    discrepancy_details = db.Column(db.Text)  # JSON string of discrepancies
    stored_bill_number = db.Column(db.String(50))
    ocr_bill_number = db.Column(db.String(50))
    stored_amount = db.Column(db.Float)
    ocr_amount = db.Column(db.Float)
    verified_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    bill = db.relationship('Bill', backref='ocr_audit_logs')

class OCRVerificationLog(db.Model):
    """Final OCR Verification Log - Tracks verification of OCR extracted data"""
    id = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('bill.id'), nullable=False)
    verified_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    status = db.Column(db.String(20), nullable=False)  # Verified, Corrected, Rejected
    mismatch_fields = db.Column(db.Text)  # JSON string of mismatched fields
    verification_date = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    remarks = db.Column(db.Text)
    
    # Relationships
    bill = db.relationship('Bill', backref='ocr_verification_logs')
    verifier = db.relationship('User', backref='verified_ocr_logs')

class OCRVendorLinkLog(db.Model):
    """Track OCR vendor matching for audit and verification"""
    id = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('bill.id'), nullable=True)
    ocr_extracted_name = db.Column(db.String(200))  # Vendor name extracted from OCR
    matched_vendor_id = db.Column(db.Integer, db.ForeignKey('vendor.id'), nullable=True)
    match_score = db.Column(db.Float)  # Similarity score (0-100)
    match_type = db.Column(db.String(20))  # exact, fuzzy, partial, no_match
    is_verified = db.Column(db.Boolean, default=False)  # Verified by Computer Organiser
    verified_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    verified_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    bill = db.relationship('Bill', backref='ocr_vendor_links')
    vendor = db.relationship('Vendor', backref='ocr_link_logs')
    verifier = db.relationship('User', foreign_keys=[verified_by], backref='verified_vendor_links')

class CreditTransaction(db.Model):
    """Credit Management System - Tracks credit transactions between vendors, customers, and company"""
    id = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('bill.id'), nullable=True)
    bill_number = db.Column(db.String(50), nullable=True)  # Store for quick reference
    vendor_id = db.Column(db.Integer, db.ForeignKey('vendor.id'), nullable=True)
    salesman_id = db.Column(db.Integer, db.ForeignKey('salesman.id'), nullable=True)
    credit_amount = db.Column(db.Numeric(10, 2), nullable=False)
    due_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='Pending', nullable=False)  # Pending, Cleared, Overdue
    payment_method = db.Column(db.String(50), default='Cash')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    vendor = db.relationship('Vendor', backref='credit_transactions')
    salesman = db.relationship('Salesman', backref='credit_transactions')
    # Note: bill relationship defined in Bill model with backref='bill_obj' to avoid conflict
    
    def update_status(self):
        """Automatically update status based on due_date"""
        today = date.today()
        if self.status == 'Cleared':
            return  # Don't change if already cleared
        if today > self.due_date:
            self.status = 'Overdue'
        elif self.status != 'Pending':
            self.status = 'Pending'
    
    def update_vendor_credits(self):
        """Update vendor credit totals when credit transaction changes"""
        if self.vendor_id:
            vendor = Vendor.query.get(self.vendor_id)
            if vendor:
                vendor.update_credit_totals()
                db.session.commit()

# Helper functions for audit logging
def get_client_ip():
    """Get client IP address from request"""
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    return request.remote_addr

def log_audit(action_type, target_type, target_id, meta=None, success=True, trace_id=None):
    """
    Log action to AuditLog
    
    Args:
        action_type: CREATE, UPDATE, DELETE, VERIFY, EXPORT, etc.
        target_type: Vendor, Bill, Credit, OCR, Delivery, etc.
        target_id: ID or unique identifier of the target
        meta: Dictionary with additional metadata (field changes, OCR scores, etc.)
        success: Boolean indicating if action succeeded
        trace_id: Optional unique trace ID for request tracking
    """
    try:
        if 'user_id' not in session:
            return None
        
        user = User.query.get(session['user_id'])
        if not user:
            return None
        
        # Generate trace_id if not provided
        if not trace_id:
            trace_id = str(uuid.uuid4())
        
        # Prepare metadata
        meta_data = meta or {}
        if isinstance(target_id, (int, str)):
            meta_data['target_id'] = str(target_id)
        
        audit_log = AuditLog(
            user_id=user.id,
            role=user.role,
            action_type=action_type.upper(),
            target_type=target_type,
            target_id=str(target_id),
            timestamp=datetime.utcnow(),
            ip_address=get_client_ip(),
            user_agent=request.headers.get('User-Agent', ''),
            meta=json.dumps(meta_data) if meta_data else None,
            success=success,
            trace_id=trace_id
        )
        
        db.session.add(audit_log)
        db.session.commit()
        return audit_log
    except Exception as e:
        print(f"Error logging audit: {e}")
        db.session.rollback()
        return None

def log_field_changes(old_obj, new_obj, exclude_fields=None):
    """Extract field changes between old and new object for audit logging"""
    if exclude_fields is None:
        exclude_fields = ['updated_at', 'created_at']
    
    changes = []
    if old_obj and new_obj:
        for column in new_obj.__table__.columns:
            field_name = column.name
            if field_name in exclude_fields:
                continue
            
            old_val = getattr(old_obj, field_name, None)
            new_val = getattr(new_obj, field_name, None)
            
            if old_val != new_val:
                changes.append({
                    'field': field_name,
                    'old': str(old_val) if old_val is not None else None,
                    'new': str(new_val) if new_val is not None else None
                })
    
    return changes

# Helper function to log activity (legacy - now uses AuditLog)
def log_activity(action_type, target_entity, target_id, old_value=None, new_value=None):
    """Log user activity (legacy function - uses AuditLog)"""
    try:
        activity = ActivityLog(
            user_id=session.get('user_id'),
            action_type=action_type,
            target_entity=target_entity,
            target_id=target_id,
            old_value=old_value,
            new_value=new_value,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        db.session.add(activity)
        db.session.commit()
    except Exception as e:
        print(f"Error logging activity: {e}")

# Authentication decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(*roles):
    """Decorator to require specific role(s)"""
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            user = User.query.get(session.get('user_id'))
            if not user or user.role not in roles:
                flash(f'Access denied. Required role: {", ".join(roles)}', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def admin_required(f):
    """Decorator for admin-only access"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def authorize(permission_code):
    """Decorator to check if user has specific permission"""
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            user = User.query.get(session.get('user_id'))
            if not user:
                flash('Please log in to access this page.', 'error')
                return redirect(url_for('login'))
            
            if not user.has_permission(permission_code):
                flash(f'Access denied. Required permission: {permission_code}', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            if not user.is_active:
                flash('Your account is disabled. Please contact administrator.', 'error')
                return render_template('login.html')
            
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            session['full_name'] = user.full_name or user.username
            
            log_activity('login', 'user', user.id, None, f"User logged in")
            log_audit('LOGIN', 'User', user.id, meta={'username': user.username, 'role': user.role})
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    if user_id:
        user = User.query.get(user_id)
        if user:
            log_audit('LOGOUT', 'User', user.id, meta={'username': user.username, 'role': user.role})
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Route to role-specific dashboards"""
    user = User.query.get(session['user_id'])
    
    # Redirect based on role
    if user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif user.role == 'delivery_man':
        return redirect(url_for('delivery_dashboard'))
    elif user.role == 'salesman':
        return redirect(url_for('salesman_dashboard'))
    elif user.role == 'computer_organiser':
        return redirect(url_for('organiser_dashboard'))
    else:
        return redirect(url_for('login'))

@app.route('/dashboard/admin')
@role_required('admin')
def admin_dashboard():
    """Admin Dashboard - Full system overview"""
    # Get all statistics
    total_invoices = Invoice.query.count()
    total_pending = Invoice.query.filter(
        Invoice.bill_amount > db.func.coalesce(
        db.select(db.func.sum(Credit.amount)).where(Credit.invoice_id == Invoice.id).scalar_subquery(), 0
        )
    ).count()
    total_paid = total_invoices - total_pending
    
    # Financial stats
    total_bill_amount = db.session.query(db.func.sum(Invoice.bill_amount)).scalar() or 0
    total_collected = db.session.query(db.func.sum(Credit.amount)).scalar() or 0
    collection_rate = (total_collected / total_bill_amount * 100) if total_bill_amount > 0 else 0
    total_outstanding = total_bill_amount - total_collected
    
    # Additional stats
    total_users = User.query.count()
    total_delivery_orders = DeliveryOrder.query.count()
    pending_deliveries = DeliveryOrder.query.filter(DeliveryOrder.status == 'pending').count()
    
    # Pending OCR verifications
    unverified_bills = Bill.query.filter(Bill.is_verified == False).count()
    
    # Recent activity
    recent_invoices = Invoice.query.order_by(Invoice.created_at.desc()).limit(5).all()
    recent_activities = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()
    
    # Get alerts
    alerts = get_alerts()
    
    return render_template('admin_dashboard.html',
                         total_invoices=total_invoices,
                         total_pending=total_pending,
                         total_paid=total_paid,
                         total_bill_amount=total_bill_amount,
                         total_collected=total_collected,
                         total_outstanding=total_outstanding,
                         collection_rate=collection_rate,
                         total_users=total_users,
                         total_delivery_orders=total_delivery_orders,
                         pending_deliveries=pending_deliveries,
                         unverified_bills=unverified_bills,
                         recent_invoices=recent_invoices,
                         recent_activities=recent_activities,
                         alerts=alerts)

@app.route('/dashboard/delivery')
@role_required('delivery_man')
def delivery_dashboard():
    """Delivery Man Dashboard - Assigned deliveries only"""
    user_id = session['user_id']
    
    # Get assigned deliveries
    assigned_deliveries = DeliveryOrder.query.filter_by(delivery_man_id=user_id).all()
    
    # Calculate stats
    total_assigned = len(assigned_deliveries)
    completed = len([d for d in assigned_deliveries if d.status == 'delivered'])
    pending = len([d for d in assigned_deliveries if d.status == 'pending'])
    in_transit = len([d for d in assigned_deliveries if d.status == 'in_transit'])
    
    completion_rate = (completed / total_assigned * 100) if total_assigned > 0 else 0
    
    # Get today's deliveries
    today = date.today()
    today_deliveries = [d for d in assigned_deliveries if d.delivery_date == today]
    
    return render_template('delivery_dashboard.html',
                         assigned_deliveries=assigned_deliveries,
                         total_assigned=total_assigned,
                         completed=completed,
                         pending=pending,
                         in_transit=in_transit,
                         completion_rate=completion_rate,
                         today_deliveries=today_deliveries)

@app.route('/dashboard/salesman')
@role_required('salesman')
def salesman_dashboard():
    """Salesman Dashboard - Sales & Billing focus"""
    user_id = session['user_id']
    
    # Get today's sales
    today = date.today()
    today_bills = Bill.query.filter(
        db.func.date(Bill.bill_date) == today,
        Bill.created_by == user_id
    ).all()
    
    # Calculate today's stats
    today_sales = sum(bill.amount for bill in today_bills)
    # Note: Credit tracking for bills would require credit transaction tracking
    # today_credits = sum(Credit.query.filter_by(invoice_id=invoice.id).all() for bill in today_bills if hasattr(bill, 'invoice'))
    
    # Get pending credits
    # pending_credits = Invoice.query.filter(Invoice.outstanding_balance > 0).all()
    
    # Recent sales
    recent_sales = Bill.query.filter_by(created_by=user_id).order_by(Bill.created_at.desc()).limit(10).all()
    
    return render_template('salesman_dashboard.html',
                         today_sales=today_sales,
                         today_bills=today_bills,
                         recent_sales=recent_sales)

@app.route('/dashboard/organiser')
@role_required('computer_organiser')
def organiser_dashboard():
    """Computer Organiser Dashboard - OCR Verification Queue"""
    # Get unverified bills (OCR queue)
    unverified_bills = Bill.query.filter_by(is_verified=False).order_by(Bill.created_at.desc()).all()
    
    # Get recently verified
    recently_verified = Bill.query.filter_by(is_verified=True).order_by(Bill.verified_at.desc()).limit(10).all()
    
    # Stats
    pending_count = len(unverified_bills)
    verified_today = Bill.query.filter(
        db.func.date(Bill.verified_at) == date.today(),
        Bill.is_verified == True
    ).count()
    
    return render_template('organiser_dashboard.html',
                         unverified_bills=unverified_bills,
                         recently_verified=recently_verified,
                         pending_count=pending_count,
                         verified_today=verified_today)

@app.route('/invoices')
@login_required
def invoices():
    page = request.args.get('page', 1, type=int)
    salesman_filter = request.args.get('salesman', type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Invoice.query
    
    if salesman_filter:
        query = query.filter(Invoice.salesman_id == salesman_filter)
    
    if status_filter == 'pending':
        query = query.filter(Invoice.bill_amount > db.func.coalesce(
            db.select(db.func.sum(Credit.amount)).where(Credit.invoice_id == Invoice.id).scalar_subquery(), 0
        ))
    elif status_filter == 'paid':
        query = query.filter(Invoice.bill_amount <= db.func.coalesce(
            db.select(db.func.sum(Credit.amount)).where(Credit.invoice_id == Invoice.id).scalar_subquery(), 0
        ))
    
    invoices = query.order_by(Invoice.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    salesmen = Salesman.query.all()
    
    return render_template('invoices.html', invoices=invoices, salesmen=salesmen)

@app.route('/invoices/new', methods=['GET', 'POST'])
@login_required
def new_invoice():
    if request.method == 'POST':
        try:
            invoice_number = request.form['invoice_number']
            salesman_id = request.form['salesman_id']
            delivery_date = datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').date()
            bill_amount = float(request.form['bill_amount'])
            
            # Check for duplicate invoice number
            existing_invoice = Invoice.query.filter_by(invoice_number=invoice_number).first()
            if existing_invoice:
                flash('Invoice number already exists!', 'error')
                return render_template('new_invoice.html', salesmen=Salesman.query.all())
            
            invoice = Invoice(
                invoice_number=invoice_number,
                salesman_id=salesman_id,
                delivery_date=delivery_date,
                bill_amount=bill_amount,
                created_by=session['user_id']
            )
            
            db.session.add(invoice)
            db.session.commit()
            
            flash('Invoice created successfully!', 'success')
            return redirect(url_for('invoices'))
        except Exception as e:
            flash(f'Error creating invoice: {str(e)}', 'error')
            return render_template('new_invoice.html', salesmen=Salesman.query.all())
    
    return render_template('new_invoice.html', salesmen=Salesman.query.all())

@app.route('/invoices/<int:invoice_id>')
@login_required
def view_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    return render_template('view_invoice.html', invoice=invoice)

@app.route('/invoices/<int:invoice_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    if request.method == 'POST':
        invoice.invoice_number = request.form['invoice_number']
        invoice.salesman_id = request.form['salesman_id']
        invoice.delivery_date = datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').date()
        invoice.bill_amount = float(request.form['bill_amount'])
        
        db.session.commit()
        flash('Invoice updated successfully!', 'success')
        return redirect(url_for('view_invoice', invoice_id=invoice.id))
    
    return render_template('edit_invoice.html', invoice=invoice, salesmen=Salesman.query.all())

@app.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
@admin_required
def delete_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    db.session.delete(invoice)
    db.session.commit()
    flash('Invoice deleted successfully!', 'success')
    return redirect(url_for('invoices'))

@app.route('/invoices/<int:invoice_id>/upload-image', methods=['GET', 'POST'])
@login_required
def upload_invoice_image(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    if request.method == 'POST':
        if 'image' not in request.files:
            flash('No image file selected.', 'error')
            return redirect(request.url)
        
        file = request.files['image']
        if file.filename == '':
            flash('No image file selected.', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            try:
                # Create upload directory if it doesn't exist
                upload_dir = app.config['UPLOAD_FOLDER']
                # For relative paths, join with root_path; for absolute paths (like /tmp), use as-is
                if not os.path.isabs(upload_dir):
                    upload_dir = os.path.join(app.root_path, upload_dir)
                os.makedirs(upload_dir, exist_ok=True)
                
                # Generate unique filename
                file_extension = file.filename.rsplit('.', 1)[1].lower()
                filename = f"invoice_{invoice_id}_{uuid.uuid4().hex}.{file_extension}"
                file_path = os.path.join(upload_dir, filename)
                
                # Save file
                file.save(file_path)
                
                # Extract text using OCR
                extracted_text, confidence, ocr_error = extract_text_from_image(file_path)
                
                if ocr_error:
                    flash(f'OCR Warning: {ocr_error}', 'warning')
                
                # Update invoice with image info
                invoice.image_filename = filename
                invoice.extracted_text = extracted_text
                invoice.ocr_confidence = confidence
                
                db.session.commit()
                
                flash(f'Image uploaded successfully! OCR confidence: {confidence:.1f}%', 'success')
                if os.environ.get('VERCEL'):
                    flash('Note: Files uploaded on Vercel are stored temporarily and may not persist. Consider using cloud storage for production.', 'info')
                return redirect(url_for('view_invoice', invoice_id=invoice_id))
            except Exception as e:
                error_msg = f'Image upload failed: {str(e)}'
                if os.environ.get('VERCEL'):
                    error_msg += ' Note: File uploads in serverless environments have limitations. Files are stored in /tmp and may not persist.'
                flash(error_msg, 'error')
                return redirect(url_for('view_invoice', invoice_id=invoice_id))
        else:
            flash('Invalid file type. Please upload a valid image file.', 'error')
    
    return render_template('upload_image.html', invoice=invoice)

@app.route('/invoices/<int:invoice_id>/image')
@login_required
def view_invoice_image(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    if not invoice.image_filename:
        flash('No image available for this invoice.', 'error')
        return redirect(url_for('view_invoice', invoice_id=invoice_id))
    
    upload_base = app.config['UPLOAD_FOLDER']
    if not os.path.isabs(upload_base):
        upload_base = os.path.join(app.root_path, upload_base)
    image_path = os.path.join(upload_base, invoice.image_filename)
    return send_file(image_path)

@app.route('/invoices/<int:invoice_id>/credits/new', methods=['GET', 'POST'])
@login_required
def new_credit(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    if request.method == 'POST':
        amount = float(request.form['amount'])
        payment_date = datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date()
        payment_method = request.form['payment_method']
        notes = request.form.get('notes', '')
        
        credit = Credit(
            invoice_id=invoice_id,
            amount=amount,
            payment_date=payment_date,
            payment_method=payment_method,
            notes=notes,
            created_by=session['user_id']
        )
        
        db.session.add(credit)
        db.session.commit()
        
        flash('Credit recorded successfully!', 'success')
        return redirect(url_for('view_invoice', invoice_id=invoice_id))
    
    return render_template('new_credit.html', invoice=invoice)

@app.route('/salesmen')
@admin_required
def salesmen():
    salesmen = Salesman.query.all()
    return render_template('salesmen.html', salesmen=salesmen)

@app.route('/salesmen/new', methods=['GET', 'POST'])
@admin_required
def new_salesman():
    if request.method == 'POST':
        salesman = Salesman(
            name=request.form['name'],
            contact=request.form['contact'],
            email=request.form.get('email', ''),
            address=request.form.get('address', '')
        )
        
        db.session.add(salesman)
        db.session.commit()
        
        flash('Salesman added successfully!', 'success')
        return redirect(url_for('salesmen'))
    
    return render_template('new_salesman.html')

@app.route('/salesmen/<int:salesman_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_salesman(salesman_id):
    salesman = Salesman.query.get_or_404(salesman_id)
    
    if request.method == 'POST':
        salesman.name = request.form['name']
        salesman.contact = request.form['contact']
        salesman.email = request.form.get('email', '')
        salesman.address = request.form.get('address', '')
        
        db.session.commit()
        flash('Salesman updated successfully!', 'success')
        return redirect(url_for('salesmen'))
    
    return render_template('edit_salesman.html', salesman=salesman)

# Vendor Management Routes
@app.route('/vendors')
@role_required('admin', 'salesman')
def vendors():
    vendors_list = Vendor.query.all()
    # Update credit totals for all vendors
    for vendor in vendors_list:
        vendor.update_credit_totals()
    db.session.commit()
    return render_template('vendors.html', vendors=vendors_list)

@app.route('/vendors/new', methods=['GET', 'POST'])
@role_required('admin', 'salesman')
def new_vendor():
    if request.method == 'POST':
        vendor = Vendor(
            name=request.form['name'],
            contact=request.form['contact'],
            email=request.form.get('email', ''),
            address=request.form.get('address', ''),
            gst_number=request.form.get('gst_number', ''),
            category=request.form.get('category', 'Supplier')
        )
        db.session.add(vendor)
        db.session.commit()
        log_activity('create', 'vendor', vendor.id, None, f"Vendor: {vendor.name}")
        flash('Vendor added successfully!', 'success')
        return redirect(url_for('vendors'))
    return render_template('new_vendor.html')

@app.route('/vendors/<int:vendor_id>/edit', methods=['GET', 'POST'])
@role_required('admin')
def edit_vendor(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    if request.method == 'POST':
        # Create a copy of old values for comparison
        old_vendor = Vendor(
            name=vendor.name,
            contact=vendor.contact,
            email=vendor.email,
            address=vendor.address,
            gst_number=vendor.gst_number,
            category=vendor.category
        )
        
        # Update vendor
        vendor.name = request.form['name']
        vendor.contact = request.form['contact']
        vendor.email = request.form.get('email', '')
        vendor.address = request.form.get('address', '')
        vendor.gst_number = request.form.get('gst_number', '')
        vendor.category = request.form.get('category', 'Supplier')
        db.session.commit()
        
        log_activity('update', 'vendor', vendor.id, old_vendor.name, vendor.name)
        
        # Log field changes
        field_changes = log_field_changes(old_vendor, vendor, exclude_fields=['updated_at', 'created_at', 'id'])
        log_audit('UPDATE', 'Vendor', vendor.id, meta={'field_changes': field_changes})
        
        flash('Vendor updated successfully!', 'success')
        return redirect(url_for('vendors'))
    return render_template('edit_vendor.html', vendor=vendor)

# Delivery Order Routes
@app.route('/deliveries')
@role_required('admin', 'delivery_man')
def deliveries():
    user_id = session['user_id']
    user = User.query.get(user_id)
    
    if user.role == 'delivery_man':
        deliveries = DeliveryOrder.query.filter_by(delivery_man_id=user_id).all()
    else:
        deliveries = DeliveryOrder.query.all()
    
    return render_template('deliveries.html', deliveries=deliveries)

@app.route('/deliveries/<int:delivery_id>')
@role_required('admin', 'delivery_man')
def view_delivery(delivery_id):
    delivery = DeliveryOrder.query.get_or_404(delivery_id)
    return render_template('view_delivery.html', delivery=delivery)

@app.route('/deliveries/<int:delivery_id>/update-status', methods=['POST'])
@role_required('admin', 'delivery_man')
def update_delivery_status(delivery_id):
    delivery = DeliveryOrder.query.get_or_404(delivery_id)
    new_status = request.form.get('status')
    delivery_remarks = request.form.get('delivery_remarks', '')
    
    old_status = delivery.status
    delivery.status = new_status
    delivery.delivery_remarks = delivery_remarks
    
    if new_status == 'delivered':
        delivery.actual_delivery_time = datetime.utcnow()
        delivery.is_confirmed = True
    
    db.session.commit()
    log_activity('update', 'delivery', delivery.id, old_status, new_status)
    log_audit('UPDATE', 'Delivery', delivery.id, meta={
        'field': 'status',
        'old': old_status,
        'new': new_status,
        'delivery_remarks': delivery_remarks
    })
    flash('Delivery status updated successfully!', 'success')
    return redirect(url_for('view_delivery', delivery_id=delivery_id))

@app.route('/deliveries/<int:delivery_id>/delete', methods=['POST'])
@admin_required
def delete_delivery(delivery_id):
    """Delete a delivery order (Admin only)"""
    delivery = DeliveryOrder.query.get_or_404(delivery_id)
    order_number = delivery.order_number
    
    try:
        log_activity('delete', 'delivery', delivery.id, f"Delivery: {order_number}", None)
        log_audit('DELETE', 'Delivery', delivery.id, meta={
            'order_number': order_number,
            'status': delivery.status
        })
        
        db.session.delete(delivery)
        db.session.commit()
        
        flash(f'Delivery order {order_number} deleted successfully.', 'success')
        return redirect(url_for('deliveries'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting delivery order: {str(e)}', 'error')
        return redirect(url_for('view_delivery', delivery_id=delivery_id))

# ============================================================================
# ENHANCED BILL MANAGEMENT ROUTES (Handbill/OCR Normal Bill)
# ============================================================================

@app.route('/bills')
@role_required('admin', 'salesman', 'computer_organiser', 'delivery_man')
def bills():
    """List all bills with filtering"""
    query = Bill.query
    
    # Filters
    bill_type = request.args.get('bill_type')
    status = request.args.get('status')
    vendor_id = request.args.get('vendor', type=int)
    payment_method = request.args.get('payment_method')
    
    if bill_type and bill_type != 'all':
        query = query.filter(Bill.bill_type == bill_type)
    if status and status != 'all':
        query = query.filter(Bill.status == status)
    if vendor_id:
        query = query.filter(Bill.vendor_id == vendor_id)
    if payment_method:
        query = query.filter(Bill.payment_method == payment_method)
    
    bills = query.order_by(Bill.created_at.desc()).all()
    vendors = Vendor.query.all()
    
    return render_template('bills.html', bills=bills, vendors=vendors)

@app.route('/bill/manual', methods=['GET', 'POST'])
@role_required('salesman', 'computer_organiser')
def create_manual_handbill():
    """Create manual handbill entry"""
    if request.method == 'POST':
        try:
            bill_number = request.form['bill_number']
            amount = float(request.form['amount'])
            bill_date = datetime.strptime(request.form['bill_date'], '%Y-%m-%d').date()
            payment_method = request.form.get('payment_method', 'Cash')
            
            # Create handbill
            bill = Bill(
                bill_number=bill_number,
                bill_type='handbill',
                amount=amount,
                bill_date=bill_date,
                payment_method=payment_method,
                vendor_id=request.form.get('vendor_id') or None,
                customer_id=request.form.get('customer_id') or None,
                salesman_id=request.form.get('salesman_id') or None,
                created_by=session['user_id'],
                is_verified=True,  # Manual bills are pre-verified
                status='verified',
                verification_status='verified'
            )
            
            db.session.add(bill)
            db.session.commit()
            
            # AUTO-CREATE CREDIT TRANSACTION if payment_method = Credit
            if payment_method == 'Credit':
                due_date_str = request.form.get('credit_due_date')
                if due_date_str:
                    try:
                        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                    except:
                        due_date = bill_date + timedelta(days=30)
                else:
                    due_date = bill_date + timedelta(days=30)
                
                credit_transaction = CreditTransaction(
                    bill_id=bill.id,
                    bill_number=bill.bill_number,
                    vendor_id=bill.vendor_id,
                    salesman_id=bill.salesman_id,
                    credit_amount=amount,
                    due_date=due_date,
                    status='Pending',
                    payment_method=payment_method
                )
                
                db.session.add(credit_transaction)
                db.session.commit()
                
                # Update vendor credit totals
                if bill.vendor_id:
                    vendor = Vendor.query.get(bill.vendor_id)
                    if vendor:
                        vendor.update_credit_totals()
                        db.session.commit()
                
                log_activity('create', 'credit_transaction', credit_transaction.id, None, 
                           f"Credit: {amount} for Handbill {bill_number}")
                log_audit('CREATE', 'Credit', credit_transaction.id, meta={
                    'bill_number': bill_number,
                    'amount': amount,
                    'vendor_id': bill.vendor_id,
                    'status': 'Pending'
                })
            
            log_activity('create', 'bill', bill.id, None, f"Handbill: {bill_number}")
            log_audit('CREATE', 'Bill', bill.id, meta={
                'bill_number': bill_number,
                'bill_type': 'handbill',
                'amount': amount,
                'payment_method': payment_method
            })
            flash('Handbill created successfully!', 'success')
            return redirect(url_for('bills'))
            
        except Exception as e:
            flash(f'Error creating handbill: {str(e)}', 'error')
    
    vendors = Vendor.query.all()
    customers = Customer.query.all()
    salesmen = Salesman.query.all()
    
    return render_template('bill_manual.html', vendors=vendors, customers=customers, salesmen=salesmen, today=date.today())

@app.route('/api/ocr-status')
@login_required
def ocr_status():
    """Check OCR engine availability"""
    status = {
        'easyocr_available': EASYOCR_AVAILABLE,
        'easyocr_initialized': _easyocr_reader is not None if EASYOCR_AVAILABLE else False,
        'message': 'EasyOCR is ready' if EASYOCR_AVAILABLE else 'EasyOCR not installed. Run: pip install easyocr'
    }
    return jsonify(status)

@app.route('/bill/upload', methods=['GET', 'POST'])
@role_required('salesman', 'computer_organiser')
def upload_ocr_bill():
    """Upload bill image and process with OCR"""
    if request.method == 'POST':
        try:
            if 'image' not in request.files:
                flash('No image file selected.', 'error')
                return redirect(url_for('upload_ocr_bill'))
            
            file = request.files['image']
            if not file or not file.filename:
                flash('No image file selected.', 'error')
                return redirect(url_for('upload_ocr_bill'))
            
            if not allowed_file(file.filename):
                flash('Invalid file type. Please upload JPEG, PNG, or PDF.', 'error')
                return redirect(url_for('upload_ocr_bill'))
            
            # Save uploaded file temporarily
            upload_base = app.config['UPLOAD_FOLDER']
            if not os.path.isabs(upload_base):
                upload_base = os.path.join(app.root_path, upload_base)
            upload_dir = os.path.join(upload_base, 'temp')
            os.makedirs(upload_dir, exist_ok=True)
            
            file_extension = file.filename.rsplit('.', 1)[1].lower()
            temp_filename = f"temp_bill_{uuid.uuid4().hex}.{file_extension}"
            temp_path = os.path.join(upload_dir, temp_filename)
            file.save(temp_path)
            
            # Extract text using OCR (may take time on first run for model download)
            flash('Processing image with OCR... This may take a few minutes on first use.', 'info')
            extracted_text, confidence, ocr_error = extract_text_from_image(temp_path)
            
            if ocr_error:
                flash(f'OCR Warning: {ocr_error}', 'warning')
            
            if not extracted_text and not ocr_error:
                flash('OCR could not extract text from the image. Please enter the data manually.', 'warning')
            
            # Parse OCR data
            existing_vendors = Vendor.query.all()
            parsed_data = parse_bill_data(extracted_text or '', existing_vendors)
            
            # Try to match vendor from OCR using fuzzy matching
            matched_vendor = None
            match_score = 0
            match_type = 'no_match'
            ocr_vendor_name = extract_vendor_name_from_text(extracted_text or '') or parsed_data.get('vendor_name')
            
            if ocr_vendor_name and existing_vendors:
                matched_vendor, match_score, match_type = match_vendor_from_ocr(
                    extracted_text or '', 
                    existing_vendors, 
                    threshold=80
                )
                if matched_vendor:
                    parsed_data['matched_vendor_id'] = matched_vendor.id
                    if match_score < 100:
                        flash(f'Vendor matched with {match_score:.0f}% confidence: {matched_vendor.name}. Please verify.', 'info')
                else:
                    flash(f'No vendor match found for: {ocr_vendor_name}. Please select manually or create new vendor.', 'warning')
            
            # Store temporary file info in session for later use
            session['temp_bill_file'] = temp_filename
            session['ocr_extracted_text'] = extracted_text
            session['ocr_confidence'] = confidence
            session['ocr_vendor_name'] = ocr_vendor_name
            session['ocr_vendor_match'] = {
                'vendor_id': matched_vendor.id if matched_vendor else None,
                'match_score': match_score,
                'match_type': match_type
            }
            session['ocr_parsed_data'] = {
                'bill_number': parsed_data['bill_number'],
                'amount': parsed_data['amount'],
                'date': parsed_data['date'].isoformat() if parsed_data['date'] else None,
                'vendor_name': parsed_data.get('vendor_name') or ocr_vendor_name
            }
            
            # Redirect to verification page
            return redirect(url_for('verify_ocr_bill'))
            
        except Exception as e:
            flash(f'Error processing bill image: {str(e)}', 'error')
    
    return render_template('bill_upload_ocr.html')

@app.route('/bill/verify', methods=['GET', 'POST'])
@role_required('salesman', 'computer_organiser')
def verify_ocr_bill():
    """Review and verify OCR extracted bill data"""
    if request.method == 'POST':
        try:
            # Get data from form
            bill_number = request.form['bill_number']
            amount = float(request.form['amount'])
            bill_date = datetime.strptime(request.form['bill_date'], '%Y-%m-%d').date()
            payment_method = request.form.get('payment_method', 'Cash')
            vendor_id = request.form.get('vendor_id') or None
            
            # Move temp file to permanent location
            temp_filename = session.get('temp_bill_file')
            if temp_filename:
                upload_base = app.config['UPLOAD_FOLDER']
                if not os.path.isabs(upload_base):
                    upload_base = os.path.join(app.root_path, upload_base)
                temp_dir = os.path.join(upload_base, 'temp')
                perm_dir = upload_base
                temp_path = os.path.join(temp_dir, temp_filename)
                
                file_extension = temp_filename.rsplit('.', 1)[1].lower()
                perm_filename = f"bill_ocr_{uuid.uuid4().hex}.{file_extension}"
                perm_path = os.path.join(perm_dir, perm_filename)
                
                # Move file
                import shutil
                shutil.move(temp_path, perm_path)
                
                extracted_text = session.get('ocr_extracted_text', '')
                confidence = session.get('ocr_confidence', 0)
                parsed_data = session.get('ocr_parsed_data', {})
            else:
                perm_filename = None
                extracted_text = ''
                confidence = 0
                parsed_data = {}
            
            # Create bill
            bill = Bill(
                bill_number=bill_number,
                bill_type='normal',
                amount=amount,
                bill_date=bill_date,
                payment_method=payment_method,
                vendor_id=vendor_id,
                salesman_id=request.form.get('salesman_id') or None,
                created_by=session['user_id'],
                image_filename=perm_filename,
                extracted_text=extracted_text,
                ocr_text=extracted_text,
                ocr_confidence=confidence,
                ocr_bill_number=parsed_data.get('bill_number'),
                ocr_amount=parsed_data.get('amount'),
                ocr_date=datetime.strptime(parsed_data['date'], '%Y-%m-%d').date() if parsed_data.get('date') and isinstance(parsed_data.get('date'), str) else None,
                ocr_vendor_name=parsed_data.get('vendor_name'),
                is_verified=False,
                status='pending',
                verification_status='unverified'
            )
            
            db.session.add(bill)
            db.session.flush()  # Get bill.id before commit
            
            # Create OCR vendor link log if vendor was matched
            ocr_vendor_match = session.get('ocr_vendor_match', {})
            ocr_vendor_name = session.get('ocr_vendor_name')
            if ocr_vendor_name:
                vendor_link_log = OCRVendorLinkLog(
                    bill_id=bill.id,
                    ocr_extracted_name=ocr_vendor_name,
                    matched_vendor_id=ocr_vendor_match.get('vendor_id'),
                    match_score=ocr_vendor_match.get('match_score', 0),
                    match_type=ocr_vendor_match.get('match_type', 'no_match'),
                    is_verified=(ocr_vendor_match.get('match_score', 0) >= 95)  # Auto-verify high confidence matches
                )
                db.session.add(vendor_link_log)
                
                # If matched vendor was not selected in form, use the matched one
                if not vendor_id and ocr_vendor_match.get('vendor_id'):
                    vendor_id = ocr_vendor_match['vendor_id']
                    bill.vendor_id = vendor_id
            
            # Create initial OCR audit log
            audit_log = OCRAuditLog(
                bill_id=bill.id,
                verification_type='initial',
                discrepancy_found=False,
                stored_bill_number=bill_number,
                ocr_bill_number=parsed_data.get('bill_number'),
                stored_amount=amount,
                ocr_amount=parsed_data.get('amount'),
                verified_by=session['user_id']
            )
            db.session.add(audit_log)
            db.session.commit()
            
            # AUTO-CREATE CREDIT TRANSACTION if payment_method = Credit
            if payment_method == 'Credit':
                due_date_str = request.form.get('credit_due_date')
                if due_date_str:
                    try:
                        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                    except:
                        due_date = bill_date + timedelta(days=30)
                else:
                    due_date = bill_date + timedelta(days=30)
                
                credit_transaction = CreditTransaction(
                    bill_id=bill.id,
                    bill_number=bill.bill_number,
                    vendor_id=bill.vendor_id,
                    salesman_id=bill.salesman_id,
                    credit_amount=amount,
                    due_date=due_date,
                    status='Pending',
                    payment_method=payment_method
                )
                
                db.session.add(credit_transaction)
                db.session.commit()
                log_activity('create', 'credit_transaction', credit_transaction.id, None, 
                           f"Credit: {amount} for OCR Bill {bill_number}")
                log_audit('CREATE', 'Credit', credit_transaction.id, meta={
                    'bill_number': bill_number,
                    'amount': amount,
                    'vendor_id': bill.vendor_id,
                    'status': 'Pending'
                })
            
            # Clear session
            session.pop('temp_bill_file', None)
            session.pop('ocr_extracted_text', None)
            session.pop('ocr_confidence', None)
            session.pop('ocr_parsed_data', None)
            
            log_activity('create', 'bill', bill.id, None, f"OCR Bill: {bill_number}")
            log_audit('CREATE', 'Bill', bill.id, meta={
                'bill_number': bill_number,
                'bill_type': 'normal',
                'amount': amount,
                'payment_method': payment_method,
                'ocr_confidence': session.get('ocr_confidence', 0)
            })
            flash('Bill saved successfully! Awaiting verification.', 'success')
            return redirect(url_for('bills'))
            
        except Exception as e:
            flash(f'Error saving bill: {str(e)}', 'error')
    
    # GET request - show verification form
    parsed_data = session.get('ocr_parsed_data', {})
    extracted_text = session.get('ocr_extracted_text', '')
    confidence = session.get('ocr_confidence', 0)
    
    vendors = Vendor.query.all()
    salesmen = Salesman.query.all()
    
    return render_template('bill_verify_ocr.html', 
                         parsed_data=parsed_data,
                         extracted_text=extracted_text,
                         confidence=confidence,
                         vendors=vendors,
                         salesmen=salesmen)

@app.route('/bill/verify/<int:bill_id>', methods=['GET', 'POST'])
@role_required('computer_organiser', 'admin')
def verify_bill_by_id(bill_id):
    """Computer Organiser verification of OCR bill"""
    bill = Bill.query.get_or_404(bill_id)
    
    if request.method == 'POST':
        try:
            # Get corrected data
            bill_number = request.form['bill_number']
            amount = float(request.form['amount'])
            bill_date = datetime.strptime(request.form['bill_date'], '%Y-%m-%d').date()
            vendor_id = request.form.get('vendor_id') or None
            
            # Store old values for comparison
            old_bill_number = bill.bill_number
            old_amount = bill.amount
            
            # Update bill
            bill.bill_number = bill_number
            bill.amount = amount
            bill.bill_date = bill_date
            bill.vendor_id = vendor_id
            bill.is_verified = True
            bill.verified_by = session['user_id']
            bill.verified_at = datetime.utcnow()
            bill.status = 'verified'
            bill.verification_status = 'verified'
            
            # Re-run OCR comparison if image exists
            if bill.image_filename and bill.extracted_text:
                re_parsed = parse_bill_data(bill.extracted_text)
                stored_data = {
                    'bill_number': bill_number,
                    'amount': amount,
                    'date': bill_date
                }
                ocr_data = {
                    'bill_number': re_parsed.get('bill_number'),
                    'amount': re_parsed.get('amount'),
                    'date': re_parsed.get('date')
                }
                
                comparison = compare_bill_data(stored_data, ocr_data)
                
                if comparison['has_discrepancy']:
                    bill.status = 'discrepancy_found'
                    bill.verification_status = 'discrepancy_found'
                    
                    # Create audit log for discrepancy
                    audit_log = OCRAuditLog(
                        bill_id=bill.id,
                        verification_type='re_verification',
                        discrepancy_found=True,
                        discrepancy_details=json.dumps(comparison['discrepancies']),
                        stored_bill_number=bill_number,
                        ocr_bill_number=ocr_data.get('bill_number'),
                        stored_amount=amount,
                        ocr_amount=ocr_data.get('amount'),
                        verified_by=session['user_id']
                    )
                    db.session.add(audit_log)
                    flash('Bill verified but discrepancies found. Please review.', 'warning')
                else:
                    # No discrepancy
                    audit_log = OCRAuditLog(
                        bill_id=bill.id,
                        verification_type='re_verification',
                        discrepancy_found=False,
                        stored_bill_number=bill_number,
                        ocr_bill_number=ocr_data.get('bill_number'),
                        stored_amount=amount,
                        ocr_amount=ocr_data.get('amount'),
                        verified_by=session['user_id']
                    )
                    db.session.add(audit_log)
            
            db.session.commit()
            log_activity('verify', 'bill', bill.id, None, f"Verified OCR Bill: {bill_number}")
            flash('Bill verified successfully!', 'success')
            return redirect(url_for('bills'))
            
        except Exception as e:
            flash(f'Error verifying bill: {str(e)}', 'error')
    
    vendors = Vendor.query.all()
    audit_logs = OCRAuditLog.query.filter_by(bill_id=bill_id).order_by(OCRAuditLog.created_at.desc()).all()
    
    return render_template('bill_verify_by_id.html', bill=bill, vendors=vendors, audit_logs=audit_logs)

@app.route('/bills/new', methods=['GET', 'POST'])
@role_required('salesman')
def new_bill():
    """Legacy route - redirects to manual handbill"""
    return redirect(url_for('create_manual_handbill'))

@app.route('/bills/<int:bill_id>')
@role_required('admin', 'salesman', 'computer_organiser')
def view_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    verified_user = None
    if bill.verified_by:
        verified_user = User.query.get(bill.verified_by)
    return render_template('view_bill.html', bill=bill, verified_user=verified_user)

@app.route('/bills/<int:bill_id>/verify', methods=['POST'])
@role_required('computer_organiser')
def verify_bill(bill_id):
    bill = Bill.query.get_or_404(bill_id)
    
    # Update bill data from verification form
    bill.bill_number = request.form.get('bill_number', bill.bill_number)
    bill.amount = float(request.form.get('amount', bill.amount))
    bill.is_verified = True
    bill.verified_by = session['user_id']
    bill.verified_at = datetime.utcnow()
    bill.status = 'verified'
    
    db.session.commit()
    log_activity('verify', 'bill', bill.id, None, f"Verified bill: {bill.bill_number}")
    flash('Bill verified successfully!', 'success')
    return redirect(url_for('organiser_dashboard'))

# ============================================================================
# CREDIT MANAGEMENT SYSTEM ROUTES
# ============================================================================

def update_overdue_credits():
    """Background task to update overdue credits automatically"""
    today = date.today()
    overdue_credits = CreditTransaction.query.filter(
        db.and_(
            CreditTransaction.status != 'Cleared',
            CreditTransaction.due_date < today
        )
    ).all()
    
    for credit in overdue_credits:
        old_status = credit.status
        credit.status = 'Overdue'
        db.session.commit()
        log_activity('update', 'credit_transaction', credit.id, old_status, 'Overdue')

@app.route('/credits')
@login_required
def credits_list():
    """List all credit transactions with filtering"""
    user = User.query.get(session['user_id'])
    
    # Role-based access control
    if user.role == 'delivery_man':
        flash('Access denied. Delivery Man role cannot view credits.', 'error')
        return redirect(url_for('dashboard'))
    
    # Base query
    query = CreditTransaction.query
    
    # Apply filters from query parameters
    vendor_id = request.args.get('vendor', type=int)
    salesman_id = request.args.get('salesman', type=int)
    status = request.args.get('status')
    payment_method = request.args.get('payment_method')
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    min_amount = request.args.get('min_amount', type=float)
    max_amount = request.args.get('max_amount', type=float)
    vendor_name = request.args.get('vendor_name')
    salesman_name = request.args.get('salesman_name')
    
    if vendor_id:
        query = query.filter(CreditTransaction.vendor_id == vendor_id)
    if salesman_id:
        query = query.filter(CreditTransaction.salesman_id == salesman_id)
    if status and status != 'all':
        query = query.filter(CreditTransaction.status == status)
    if payment_method:
        query = query.filter(CreditTransaction.payment_method == payment_method)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(CreditTransaction.due_date >= start)
        except:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(CreditTransaction.due_date <= end)
        except:
            pass
    if min_amount:
        query = query.filter(CreditTransaction.credit_amount >= min_amount)
    if max_amount:
        query = query.filter(CreditTransaction.credit_amount <= max_amount)
    if vendor_name:
        query = query.join(Vendor).filter(Vendor.name.ilike(f'%{vendor_name}%'))
    if salesman_name:
        query = query.join(Salesman).filter(Salesman.name.ilike(f'%{salesman_name}%'))
    
    # Update status before displaying
    update_overdue_credits()
    
    credits = query.order_by(CreditTransaction.created_at.desc()).all()
    
    vendors = Vendor.query.all()
    salesmen = Salesman.query.all()
    
    return render_template('credits_list.html', 
                         credits=credits,
                         vendors=vendors,
                         salesmen=salesmen,
                         today=date.today(),
                         current_filters={
                             'vendor': vendor_id,
                             'salesman': salesman_id,
                             'status': status,
                             'payment_method': payment_method,
                             'start': start_date,
                             'end': end_date,
                             'min_amount': min_amount,
                             'max_amount': max_amount,
                             'vendor_name': vendor_name,
                             'salesman_name': salesman_name
                         })

@app.route('/credits/dashboard')
@login_required
def credits_dashboard():
    """Summary dashboard with analytics"""
    user = User.query.get(session['user_id'])
    
    if user.role == 'delivery_man':
        flash('Access denied. Delivery Man role cannot view credits.', 'error')
        return redirect(url_for('dashboard'))
    
    # Update overdue credits first
    update_overdue_credits()
    
    # Summary statistics
    total_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).scalar() or 0
    pending_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        CreditTransaction.status == 'Pending'
    ).scalar() or 0
    cleared_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        CreditTransaction.status == 'Cleared'
    ).scalar() or 0
    overdue_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        CreditTransaction.status == 'Overdue'
    ).scalar() or 0
    
    # Vendor-wise breakdown
    vendor_breakdown = db.session.query(
        Vendor.name,
        db.func.sum(CreditTransaction.credit_amount).label('total'),
        db.func.count(CreditTransaction.id).label('count')
    ).join(
        CreditTransaction, Vendor.id == CreditTransaction.vendor_id
    ).group_by(Vendor.id, Vendor.name).all()
    
    # Upcoming due alerts (next 5 due credits within 3 days)
    three_days_from_now = date.today() + timedelta(days=3)
    upcoming_due = CreditTransaction.query.filter(
        db.and_(
            CreditTransaction.status != 'Cleared',
            CreditTransaction.due_date <= three_days_from_now,
            CreditTransaction.due_date >= date.today()
        )
    ).order_by(CreditTransaction.due_date.asc()).limit(5).all()
    
    # Overdue alerts
    overdue_list = CreditTransaction.query.filter(
        CreditTransaction.status == 'Overdue'
    ).order_by(CreditTransaction.due_date.asc()).limit(10).all()
    
    return render_template('credits_dashboard.html',
                         total_credits=float(total_credits),
                         pending_credits=float(pending_credits),
                         cleared_credits=float(cleared_credits),
                         overdue_credits=float(overdue_credits),
                         vendor_breakdown=vendor_breakdown,
                         upcoming_due=upcoming_due,
                         overdue_list=overdue_list)

@app.route('/credits/<int:credit_id>')
@login_required
def view_credit(credit_id):
    """View a single credit transaction"""
    credit = CreditTransaction.query.get_or_404(credit_id)
    
    # Check access
    user = User.query.get(session['user_id'])
    if user.role == 'delivery_man':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('view_credit.html', credit=credit, today=date.today())

@app.route('/credits/<int:credit_id>/update-status', methods=['POST'])
@role_required('admin')
def update_credit_status(credit_id):
    """Update credit status (Admin only, can override)"""
    credit = CreditTransaction.query.get_or_404(credit_id)
    old_status = credit.status
    new_status = request.form.get('status')
    
    if new_status in ['Pending', 'Cleared', 'Overdue']:
        credit.status = new_status
        credit.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Update vendor credit totals
        credit.update_vendor_credits()
        
        log_activity('update', 'credit_transaction', credit.id, old_status, new_status)
        log_audit('UPDATE', 'Credit', credit.id, meta={
            'field': 'status',
            'old': old_status,
            'new': new_status,
            'updated_by': session.get('user_id')
        })
        flash(f'Credit status updated to {new_status}', 'success')
    else:
        flash('Invalid status', 'error')
    
    return redirect(url_for('view_credit', credit_id=credit_id))

@app.route('/credits/api/summary')
@login_required
def credits_api_summary():
    """API endpoint for dashboard summary data"""
    user = User.query.get(session['user_id'])
    
    if user.role == 'delivery_man':
        return jsonify({'error': 'Access denied'}), 403
    
    update_overdue_credits()
    
    total_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).scalar() or 0
    pending_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        CreditTransaction.status == 'Pending'
    ).scalar() or 0
    cleared_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        CreditTransaction.status == 'Cleared'
    ).scalar() or 0
    overdue_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        CreditTransaction.status == 'Overdue'
    ).scalar() or 0
    
    vendor_breakdown = db.session.query(
        Vendor.name,
        db.func.sum(CreditTransaction.credit_amount).label('total')
    ).join(
        CreditTransaction, Vendor.id == CreditTransaction.vendor_id
    ).group_by(Vendor.id, Vendor.name).all()
    
    return jsonify({
        'total_credits': float(total_credits),
        'pending_credits': float(pending_credits),
        'cleared_credits': float(cleared_credits),
        'overdue_credits': float(overdue_credits),
        'vendor_breakdown': [{'name': v[0], 'total': float(v[1])} for v in vendor_breakdown]
    })

# Customer Management Routes
@app.route('/database/structure')
@admin_required
def database_structure():
    """View database structure and all tables (Admin only)"""
    inspector = inspect(db.engine)
    tables_info = {}
    
    # Get all table names
    table_names = inspector.get_table_names()
    
    for table_name in table_names:
        # Get columns
        columns_info = []
        for column in inspector.get_columns(table_name):
            col_info = {
                'name': column['name'],
                'type': str(column['type']),
                'nullable': column['nullable'],
                'default': str(column.get('default', '')) if column.get('default') is not None else None,
                'primary_key': False,
                'autoincrement': column.get('autoincrement', False),
                'foreign_key': None
            }
            columns_info.append(col_info)
        
        # Get primary keys
        pk_constraint = inspector.get_pk_constraint(table_name)
        if pk_constraint and 'constrained_columns' in pk_constraint:
            for pk_col in pk_constraint['constrained_columns']:
                for col in columns_info:
                    if col['name'] == pk_col:
                        col['primary_key'] = True
        
        # Get foreign keys
        foreign_keys = []
        fk_constraints = inspector.get_foreign_keys(table_name)
        for fk in fk_constraints:
            fk_info = {
                'column': fk['constrained_columns'][0] if fk['constrained_columns'] else None,
                'ref_table': fk['referred_table'],
                'ref_column': fk['referred_columns'][0] if fk['referred_columns'] else None
            }
            foreign_keys.append(fk_info)
            
            # Mark column as foreign key
            for col in columns_info:
                if col['name'] == fk_info['column']:
                    col['foreign_key'] = f"{fk_info['ref_table']}.{fk_info['ref_column']}"
        
        # Get indexes
        indexes_info = []
        indexes = inspector.get_indexes(table_name)
        for idx in indexes:
            indexes_info.append({
                'name': idx['name'],
                'columns': idx['column_names'],
                'unique': idx['unique']
            })
        
        # Get row count
        row_count = None
        try:
            result = db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}"))
            row_count = result.scalar()
        except:
            pass
        
        tables_info[table_name] = {
            'columns': columns_info,
            'foreign_keys': foreign_keys,
            'indexes': indexes_info,
            'row_count': row_count
        }
    
    # Get database info
    db_info = {
        'type': 'SQLite',
        'path': None,
        'size': None
    }
    
    # Try to get database path
    db_uri = app.config['SQLALCHEMY_DATABASE_URI']
    if db_uri.startswith('sqlite:///'):
        db_path = db_uri.replace('sqlite:///', '')
        if os.path.exists(db_path):
            db_info['path'] = os.path.abspath(db_path)
            try:
                size_bytes = os.path.getsize(db_path)
                if size_bytes < 1024:
                    db_info['size'] = f"{size_bytes} B"
                elif size_bytes < 1024 * 1024:
                    db_info['size'] = f"{size_bytes / 1024:.2f} KB"
                else:
                    db_info['size'] = f"{size_bytes / (1024 * 1024):.2f} MB"
            except:
                pass
    
    return render_template('database_structure.html', 
                         tables=tables_info, 
                         db_info=db_info)

# RBAC and Audit Log Routes
@app.route('/roles/list')
@admin_required
def roles_list():
    """View all roles and their permissions"""
    roles = Role.query.all()
    permissions = Permission.query.all()
    return render_template('roles_list.html', roles=roles, permissions=permissions)

@app.route('/role/assign', methods=['POST'])
@admin_required
def assign_role_permission():
    """Assign permission to role"""
    role_id = request.form.get('role_id', type=int)
    permission_id = request.form.get('permission_id', type=int)
    
    role = Role.query.get_or_404(role_id)
    permission = Permission.query.get_or_404(permission_id)
    
    if permission not in role.permissions.all():
        role.permissions.append(permission)
        db.session.commit()
        log_audit('UPDATE', 'Role', role.id, meta={'action': 'assign_permission', 'permission': permission.code})
        flash(f'Permission {permission.code} assigned to role {role.name}', 'success')
    else:
        flash('Permission already assigned', 'info')
    
    return redirect(url_for('roles_list'))

@app.route('/logs')
@admin_required
def audit_logs():
    """View audit logs with filtering"""
    # Get filter parameters
    user_id = request.args.get('user_id', type=int)
    role_filter = request.args.get('role')
    action_type = request.args.get('action_type')
    target_type = request.args.get('target_type')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    success_filter = request.args.get('success')
    
    # Base query
    query = AuditLog.query
    
    # Apply filters
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    if role_filter:
        query = query.filter(AuditLog.role == role_filter)
    if action_type:
        query = query.filter(AuditLog.action_type == action_type.upper())
    if target_type:
        query = query.filter(AuditLog.target_type == target_type)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.timestamp >= date_from_obj)
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.timestamp < date_to_obj)
        except:
            pass
    if success_filter is not None:
        query = query.filter(AuditLog.success == (success_filter.lower() == 'true'))
    
    # Order by timestamp descending
    query = query.order_by(AuditLog.timestamp.desc())
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    logs = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get filter options
    users = User.query.all()
    roles = list(ROLES.values())
    action_types = ['CREATE', 'UPDATE', 'DELETE', 'VERIFY', 'EXPORT', 'LOGIN', 'LOGOUT']
    target_types = ['Vendor', 'Bill', 'Credit', 'OCR', 'Delivery', 'User', 'Invoice']
    
    return render_template('audit_logs.html',
                         logs=logs,
                         users=users,
                         roles=roles,
                         action_types=action_types,
                         target_types=target_types)

@app.route('/logs/<int:log_id>')
@admin_required
def view_audit_log(log_id):
    """View detailed audit log entry"""
    log = AuditLog.query.get_or_404(log_id)
    meta = log.get_meta()
    return render_template('view_audit_log.html', log=log, meta=meta)

@app.route('/logs/export')
@authorize('reports.export')
def export_audit_logs():
    """Export audit logs to CSV"""
    # Apply same filters as /logs
    user_id = request.args.get('user_id', type=int)
    role_filter = request.args.get('role')
    action_type = request.args.get('action_type')
    target_type = request.args.get('target_type')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = AuditLog.query
    
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    if role_filter:
        query = query.filter(AuditLog.role == role_filter)
    if action_type:
        query = query.filter(AuditLog.action_type == action_type.upper())
    if target_type:
        query = query.filter(AuditLog.target_type == target_type)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.timestamp >= date_from_obj)
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.timestamp < date_to_obj)
        except:
            pass
    
    logs = query.order_by(AuditLog.timestamp.desc()).limit(10000).all()
    
    # Create CSV
    output = io.StringIO()
    output.write('ID,Timestamp,User,Role,Action,Target Type,Target ID,IP Address,Success,Trace ID\n')
    
    for log in logs:
        user_name = log.user.username if log.user else 'Unknown'
        output.write(f'{log.id},{log.timestamp},{user_name},{log.role},{log.action_type},'
                    f'{log.target_type},{log.target_id},{log.ip_address or ""},'
                    f'{log.success},{log.trace_id}\n')
    
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=audit_logs_{datetime.now().strftime("%Y%m%d")}.csv'
    
    log_audit('EXPORT', 'AuditLog', 'all', meta={'format': 'CSV', 'count': len(logs)})
    
    return response

@app.route('/entity/<target_type>/<target_id>/history')
@login_required
def entity_history(target_type, target_id):
    """Get action timeline for a specific entity"""
    logs = AuditLog.query.filter(
        db.and_(
            AuditLog.target_type == target_type,
            AuditLog.target_id == str(target_id)
        )
    ).order_by(AuditLog.timestamp.desc()).limit(100).all()
    
    return render_template('entity_history.html', logs=logs, target_type=target_type, target_id=target_id)

@app.route('/customers')
@role_required('admin', 'salesman')
def customers():
    customers = Customer.query.all()
    return render_template('customers.html', customers=customers)

@app.route('/customers/new', methods=['GET', 'POST'])
@admin_required
def new_customer():
    if request.method == 'POST':
        customer = Customer(
            name=request.form['name'],
            contact=request.form['contact'],
            email=request.form.get('email', ''),
            address=request.form.get('address', ''),
            company_name=request.form.get('company_name', ''),
            gst_number=request.form.get('gst_number', ''),
            credit_limit=float(request.form.get('credit_limit', 0))
        )
        
        db.session.add(customer)
        db.session.commit()
        
        flash('Customer added successfully!', 'success')
        return redirect(url_for('customers'))
    
    return render_template('new_customer.html')

@app.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    
    if request.method == 'POST':
        customer.name = request.form['name']
        customer.contact = request.form['contact']
        customer.email = request.form.get('email', '')
        customer.address = request.form.get('address', '')
        customer.company_name = request.form.get('company_name', '')
        customer.gst_number = request.form.get('gst_number', '')
        customer.credit_limit = float(request.form.get('credit_limit', 0))
        
        db.session.commit()
        flash('Customer updated successfully!', 'success')
        return redirect(url_for('customers'))
    
    return render_template('edit_customer.html', customer=customer)

# User Management Routes
@app.route('/users')
@admin_required
def users():
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/users/new', methods=['GET', 'POST'])
@admin_required
def new_user():
    if request.method == 'POST':
        user = User(
            username=request.form['username'],
            email=request.form['email'],
            role=request.form['role'],
            full_name=request.form.get('full_name', ''),
            phone=request.form.get('phone', ''),
            is_active=True
        )
        user.set_password(request.form['password'])
        
        db.session.add(user)
        db.session.commit()
        
        log_activity('create', 'user', user.id, None, f"User: {user.username}")
        flash('User added successfully!', 'success')
        return redirect(url_for('users'))
    
    return render_template('new_user.html', roles=ROLES)

@app.route('/deliveries/new', methods=['GET', 'POST'])
@role_required('admin')
def new_delivery():
    if request.method == 'POST':
        # Generate unique order number
        order_number = f"DO{datetime.now().strftime('%Y%m%d')}{DeliveryOrder.query.count() + 1}"
        
        delivery = DeliveryOrder(
            order_number=order_number,
            delivery_man_id=int(request.form['delivery_man_id']),
            invoice_id=int(request.form['invoice_id']) if request.form.get('invoice_id') else None,
            delivery_address=request.form['delivery_address'],
            customer_contact=request.form.get('customer_contact', ''),
            delivery_date=datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').date(),
            status='pending'
        )
        
        db.session.add(delivery)
        db.session.commit()
        
        log_activity('create', 'delivery', delivery.id, None, f"Delivery: {order_number}")
        flash('Delivery order created successfully!', 'success')
        return redirect(url_for('deliveries'))
    
    delivery_men = User.query.filter_by(role='delivery_man').all()
    invoices = Invoice.query.all()
    
    return render_template('new_delivery.html', delivery_men=delivery_men, invoices=invoices)

# Export Routes
@app.route('/export/pdf')
@login_required
def export_pdf():
    salesman_filter = request.args.get('salesman', type=int)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = Invoice.query
    
    if salesman_filter:
        query = query.filter(Invoice.salesman_id == salesman_filter)
    
    if date_from:
        query = query.filter(Invoice.delivery_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(Invoice.delivery_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    invoices = query.order_by(Invoice.delivery_date.desc()).all()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    title = Paragraph("Credit Management Report - Skanda Enterprises", title_style)
    
    # Report details
    report_date = datetime.now().strftime("%B %d, %Y")
    report_info = Paragraph(f"Generated on: {report_date}", styles['Normal'])
    
    # Summary statistics
    total_invoices = len(invoices)
    total_amount = sum(invoice.bill_amount for invoice in invoices)
    total_collected = sum(invoice.total_credits for invoice in invoices)
    collection_rate = (total_collected / total_amount * 100) if total_amount > 0 else 0
    
    summary_data = [
        ['Total Invoices', str(total_invoices)],
        ['Total Amount', f'₹{total_amount:,.2f}'],
        ['Total Collected', f'₹{total_collected:,.2f}'],
        ['Collection Rate', f'{collection_rate:.1f}%']
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    # Invoice details table
    if invoices:
        invoice_data = [['Invoice #', 'Customer', 'Salesman', 'Date', 'Amount', 'Collected', 'Outstanding', 'Status']]
        
        for invoice in invoices:
            customer_name = invoice.customer.name if invoice.customer else 'N/A'
            status = 'Paid' if invoice.is_paid else 'Pending'
            invoice_data.append([
                invoice.invoice_number,
                customer_name,
                invoice.salesman.name,
                invoice.delivery_date.strftime('%Y-%m-%d'),
                f'₹{invoice.bill_amount:,.2f}',
                f'₹{invoice.total_credits:,.2f}',
                f'₹{invoice.outstanding_balance:,.2f}',
                status
            ])
        
        invoice_table = Table(invoice_data, colWidths=[1*inch, 1.2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch])
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
    else:
        invoice_table = Paragraph("No invoices found for the selected criteria.", styles['Normal'])
    
    # Build PDF
    elements = [title, Spacer(1, 20), report_info, Spacer(1, 20), 
                Paragraph("Summary", styles['Heading2']), summary_table, Spacer(1, 20),
                Paragraph("Invoice Details", styles['Heading2']), invoice_table]
    
    doc.build(elements)
    
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=credit_report_{datetime.now().strftime("%Y%m%d")}.pdf'
    
    return response

@app.route('/export/excel')
@login_required
def export_excel():
    salesman_filter = request.args.get('salesman', type=int)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = Invoice.query
    
    if salesman_filter:
        query = query.filter(Invoice.salesman_id == salesman_filter)
    
    if date_from:
        query = query.filter(Invoice.delivery_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(Invoice.delivery_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    invoices = query.order_by(Invoice.delivery_date.desc()).all()
    
    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Credit Report"
    
    # Headers
    headers = ['Invoice #', 'Customer', 'Salesman', 'Date', 'Amount', 'Collected', 'Outstanding', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, invoice in enumerate(invoices, 2):
        customer_name = invoice.customer.name if invoice.customer else 'N/A'
        status = 'Paid' if invoice.is_paid else 'Pending'
        
        ws.cell(row=row, column=1, value=invoice.invoice_number)
        ws.cell(row=row, column=2, value=customer_name)
        ws.cell(row=row, column=3, value=invoice.salesman.name)
        ws.cell(row=row, column=4, value=invoice.delivery_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=invoice.bill_amount)
        ws.cell(row=row, column=6, value=invoice.total_credits)
        ws.cell(row=row, column=7, value=invoice.outstanding_balance)
        ws.cell(row=row, column=8, value=status)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=credit_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response

# OCR Verification Helper Functions
def detect_ocr_mismatches(bill):
    """Detect mismatches between OCR extracted data and manually entered data"""
    mismatches = {}
    
    if bill.bill_type != 'normal':
        return mismatches  # Only check OCR bills
    
    # Check bill number mismatch
    if hasattr(bill, 'ocr_bill_number') and bill.ocr_bill_number:
        if bill.bill_number != bill.ocr_bill_number:
            mismatches['bill_number'] = {
                'ocr': bill.ocr_bill_number,
                'stored': bill.bill_number
            }
    
    # Check amount mismatch (allow 5 rupee difference)
    if hasattr(bill, 'ocr_amount') and bill.ocr_amount:
        amount_diff = abs(float(bill.amount) - float(bill.ocr_amount))
        if amount_diff > 5:
            mismatches['amount'] = {
                'ocr': bill.ocr_amount,
                'stored': bill.amount,
                'difference': amount_diff
            }
    
    # Check date mismatch
    if hasattr(bill, 'ocr_date') and bill.ocr_date:
        if isinstance(bill.ocr_date, str):
            try:
                ocr_date = datetime.strptime(bill.ocr_date, '%Y-%m-%d').date()
            except:
                ocr_date = None
        else:
            ocr_date = bill.ocr_date
        
        if ocr_date and bill.bill_date != ocr_date:
            mismatches['date'] = {
                'ocr': str(ocr_date),
                'stored': str(bill.bill_date)
            }
    
    # Check vendor mismatch (if OCR vendor name exists)
    ocr_vendor_link = OCRVendorLinkLog.query.filter_by(bill_id=bill.id).first()
    if ocr_vendor_link and ocr_vendor_link.ocr_extracted_name:
        if bill.vendor and bill.vendor.name.lower() != ocr_vendor_link.ocr_extracted_name.lower():
            mismatches['vendor'] = {
                'ocr': ocr_vendor_link.ocr_extracted_name,
                'stored': bill.vendor.name if bill.vendor else 'None'
            }
    
    return mismatches

def run_daily_ocr_verification():
    """Run end-of-day OCR verification on all unverified OCR bills"""
    flagged_bills = []
    
    # Get all OCR bills that haven't been verified yet
    ocr_bills = Bill.query.filter(
        Bill.bill_type == 'normal',
        Bill.verification_status != 'verified'
    ).all()
    
    for bill in ocr_bills:
        mismatches = detect_ocr_mismatches(bill)
        if mismatches:
            flagged_bills.append({
                'bill': bill,
                'mismatches': mismatches
            })
    
    return flagged_bills

def get_alerts():
    """Get all system alerts (overdue credits, upcoming due, OCR mismatches)"""
    alerts = {
        'overdue_credits': [],
        'overdue_total': 0,
        'upcoming_due': [],
        'ocr_mismatches': 0
    }
    
    # Overdue credits
    overdue_credits = CreditTransaction.query.filter(
        CreditTransaction.status == 'Overdue'
    ).order_by(CreditTransaction.due_date.asc()).limit(10).all()
    alerts['overdue_credits'] = overdue_credits
    alerts['overdue_total'] = sum(float(c.credit_amount) for c in overdue_credits)
    
    # Upcoming due (within 3 days)
    three_days_from_now = date.today() + timedelta(days=3)
    upcoming_due = CreditTransaction.query.filter(
        db.and_(
            CreditTransaction.status != 'Cleared',
            CreditTransaction.due_date <= three_days_from_now,
            CreditTransaction.due_date >= date.today()
        )
    ).order_by(CreditTransaction.due_date.asc()).limit(5).all()
    alerts['upcoming_due'] = upcoming_due
    
    # OCR mismatches count
    flagged_bills = run_daily_ocr_verification()
    alerts['ocr_mismatches'] = len(flagged_bills)
    
    return alerts

@app.route('/reports')
@login_required
def reports():
    salesman_filter = request.args.get('salesman', type=int)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = Invoice.query
    
    if salesman_filter:
        query = query.filter(Invoice.salesman_id == salesman_filter)
    
    if date_from:
        query = query.filter(Invoice.delivery_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(Invoice.delivery_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    invoices = query.order_by(Invoice.delivery_date.desc()).all()
    salesmen = Salesman.query.all()
    
    # Calculate collection rate for filtered data
    total_bill_amount = sum(invoice.bill_amount for invoice in invoices)
    total_collected = sum(invoice.total_credits for invoice in invoices)
    collection_rate = (total_collected / total_bill_amount * 100) if total_bill_amount > 0 else 0
    
    return render_template('reports.html', 
                         invoices=invoices, 
                         salesmen=salesmen,
                         total_bill_amount=total_bill_amount,
                         total_collected=total_collected,
                         collection_rate=collection_rate)

# Reports & Analytics Routes
@app.route('/reports/analytics')
@login_required
def analytics_dashboard():
    """Main analytics dashboard with key metrics and charts"""
    user = User.query.get(session['user_id'])
    
    # Get filter parameters
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    vendor_id = request.args.get('vendor_id', type=int)
    salesman_id = request.args.get('salesman_id', type=int)
    
    # Default to current month if no date range
    if not date_from or not date_to:
        today = date.today()
        date_from = date(today.year, today.month, 1).isoformat()
        date_to = today.isoformat()
    
    # Convert to date objects
    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Base queries with date filter
    bills_query = Bill.query.filter(
        db.and_(
            db.func.date(Bill.bill_date) >= date_from_obj,
            db.func.date(Bill.bill_date) <= date_to_obj
        )
    )
    
    credits_query = CreditTransaction.query.filter(
        db.and_(
            db.func.date(CreditTransaction.created_at) >= date_from_obj,
            db.func.date(CreditTransaction.created_at) <= date_to_obj
        )
    )
    
    # Apply additional filters
    if vendor_id:
        bills_query = bills_query.filter(Bill.vendor_id == vendor_id)
        credits_query = credits_query.filter(CreditTransaction.vendor_id == vendor_id)
    
    if salesman_id:
        bills_query = bills_query.filter(Bill.salesman_id == salesman_id)
        credits_query = credits_query.filter(CreditTransaction.salesman_id == salesman_id)
    
    # Calculate metrics
    total_sales = db.session.query(db.func.sum(Bill.amount)).filter(
        bills_query.whereclause
    ).scalar() or 0
    
    total_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        credits_query.whereclause
    ).scalar() or 0
    
    pending_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        credits_query.whereclause,
        CreditTransaction.status == 'Pending'
    ).scalar() or 0
    
    cleared_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        credits_query.whereclause,
        CreditTransaction.status == 'Cleared'
    ).scalar() or 0
    
    overdue_credits = db.session.query(db.func.sum(CreditTransaction.credit_amount)).filter(
        credits_query.whereclause,
        CreditTransaction.status == 'Overdue'
    ).scalar() or 0
    
    # Vendor Performance (Top 5 by transaction volume)
    vendor_performance = db.session.query(
        Vendor.name,
        db.func.count(Bill.id).label('bill_count'),
        db.func.sum(Bill.amount).label('total_amount')
    ).join(Bill, Vendor.id == Bill.vendor_id).filter(
        bills_query.whereclause
    ).group_by(Vendor.id, Vendor.name).order_by(
        db.func.sum(Bill.amount).desc()
    ).limit(5).all()
    
    # Salesman Performance
    salesman_performance = db.session.query(
        Salesman.name,
        db.func.count(Bill.id).label('bill_count'),
        db.func.sum(Bill.amount).label('total_sales')
    ).join(Bill, Salesman.id == Bill.salesman_id).filter(
        bills_query.whereclause
    ).group_by(Salesman.id, Salesman.name).all()
    
    # Delivery Efficiency
    deliveries_query = DeliveryOrder.query.filter(
        db.and_(
            db.func.date(DeliveryOrder.delivery_date) >= date_from_obj,
            db.func.date(DeliveryOrder.delivery_date) <= date_to_obj
        )
    )
    
    total_deliveries = deliveries_query.count()
    delivered_on_time = deliveries_query.filter(
        DeliveryOrder.status == 'delivered',
        DeliveryOrder.actual_delivery_time <= db.func.datetime(DeliveryOrder.delivery_date, '+1 day')
    ).count()
    
    delivery_efficiency = (delivered_on_time / total_deliveries * 100) if total_deliveries > 0 else 0
    
    # Credit Status by Vendor (for chart)
    credit_by_vendor = db.session.query(
        Vendor.name,
        CreditTransaction.status,
        db.func.sum(CreditTransaction.credit_amount).label('total')
    ).join(CreditTransaction, Vendor.id == CreditTransaction.vendor_id).filter(
        credits_query.whereclause
    ).group_by(Vendor.name, CreditTransaction.status).all()
    
    # Vendor Category Distribution
    vendor_categories = db.session.query(
        Vendor.category,
        db.func.count(Bill.id).label('count')
    ).join(Bill, Vendor.id == Bill.vendor_id).filter(
        bills_query.whereclause
    ).group_by(Vendor.category).all()
    
    # Sales Trend (by day)
    sales_trend = db.session.query(
        db.func.date(Bill.bill_date).label('date'),
        db.func.sum(Bill.amount).label('total')
    ).filter(
        bills_query.whereclause
    ).group_by(db.func.date(Bill.bill_date)).order_by('date').all()
    
    return render_template('analytics_dashboard.html',
                         total_sales=float(total_sales),
                         total_credits=float(total_credits),
                         pending_credits=float(pending_credits),
                         cleared_credits=float(cleared_credits),
                         overdue_credits=float(overdue_credits),
                         vendor_performance=vendor_performance,
                         salesman_performance=salesman_performance,
                         delivery_efficiency=delivery_efficiency,
                         total_deliveries=total_deliveries,
                         credit_by_vendor=credit_by_vendor,
                         vendor_categories=vendor_categories,
                         sales_trend=sales_trend,
                         vendors=Vendor.query.all(),
                         salesmen=Salesman.query.all(),
                         date_from=date_from,
                         date_to=date_to,
                         selected_vendor=vendor_id,
                         selected_salesman=salesman_id)

@app.route('/reports/credit-summary')
@login_required
def credit_summary_report():
    """Credit summary report API endpoint"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    vendor_id = request.args.get('vendor_id', type=int)
    status = request.args.get('status')
    
    query = CreditTransaction.query
    
    if date_from:
        query = query.filter(CreditTransaction.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(CreditTransaction.created_at <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
    if vendor_id:
        query = query.filter(CreditTransaction.vendor_id == vendor_id)
    if status:
        query = query.filter(CreditTransaction.status == status)
    
    credits = query.all()
    
    return jsonify({
        'total': len(credits),
        'total_amount': sum(float(c.credit_amount) for c in credits),
        'pending': sum(float(c.credit_amount) for c in credits if c.status == 'Pending'),
        'cleared': sum(float(c.credit_amount) for c in credits if c.status == 'Cleared'),
        'overdue': sum(float(c.credit_amount) for c in credits if c.status == 'Overdue'),
        'credits': [{
            'id': c.id,
            'bill_number': c.bill_number,
            'vendor': c.vendor.name if c.vendor else 'N/A',
            'amount': float(c.credit_amount),
            'status': c.status,
            'due_date': c.due_date.isoformat() if c.due_date else None
        } for c in credits]
    })

@app.route('/reports/sales-performance')
@login_required
def sales_performance_report():
    """Sales performance report API endpoint"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    salesman_id = request.args.get('salesman_id', type=int)
    
    query = Bill.query
    
    if date_from:
        query = query.filter(Bill.bill_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(Bill.bill_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if salesman_id:
        query = query.filter(Bill.salesman_id == salesman_id)
    
    bills = query.all()
    
    return jsonify({
        'total_bills': len(bills),
        'total_sales': sum(float(b.amount) for b in bills),
        'handbill_count': len([b for b in bills if b.bill_type == 'handbill']),
        'ocr_count': len([b for b in bills if b.bill_type == 'normal']),
        'bills': [{
            'id': b.id,
            'bill_number': b.bill_number,
            'type': b.bill_type,
            'amount': float(b.amount),
            'date': b.bill_date.isoformat() if b.bill_date else None
        } for b in bills]
    })

@app.route('/reports/vendor-performance')
@login_required
def vendor_performance_report():
    """Vendor performance report API endpoint"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = Vendor.query
    
    vendors = query.all()
    result = []
    
    for vendor in vendors:
        bills_query = Bill.query.filter(Bill.vendor_id == vendor.id)
        if date_from:
            bills_query = bills_query.filter(Bill.bill_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        if date_to:
            bills_query = bills_query.filter(Bill.bill_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        
        bills = bills_query.all()
        total_sales = sum(float(b.amount) for b in bills)
        
        result.append({
            'vendor': vendor.name,
            'total_transactions': len(bills),
            'total_sales': total_sales,
            'outstanding_credit': float(vendor.outstanding_credit) if vendor.outstanding_credit else 0
        })
    
    result.sort(key=lambda x: x['total_sales'], reverse=True)
    
    return jsonify({'vendors': result})

@app.route('/reports/export')
@authorize('reports.export')
def export_reports():
    """Export reports to Excel or PDF"""
    export_type = request.args.get('type', 'excel')  # excel or pdf
    report_type = request.args.get('report')  # credit, sales, vendor, delivery
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    vendor_id = request.args.get('vendor_id', type=int)
    salesman_id = request.args.get('salesman_id', type=int)
    
    if export_type == 'excel':
        # Create Excel file using openpyxl or csv
        output = io.StringIO()
        output.write('Report Type,Date From,Date To\n')
        output.write(f'{report_type},{date_from},{date_to}\n\n')
        
        if report_type == 'credit':
            query = CreditTransaction.query
            if date_from:
                query = query.filter(CreditTransaction.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
            if date_to:
                query = query.filter(CreditTransaction.created_at <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            if vendor_id:
                query = query.filter(CreditTransaction.vendor_id == vendor_id)
            
            credits = query.all()
            output.write('Bill Number,Vendor,Amount,Status,Due Date\n')
            for c in credits:
                output.write(f'{c.bill_number},{c.vendor.name if c.vendor else "N/A"},{c.credit_amount},{c.status},{c.due_date}\n')
        
        elif report_type == 'sales':
            query = Bill.query
            if date_from:
                query = query.filter(Bill.bill_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
            if date_to:
                query = query.filter(Bill.bill_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
            if salesman_id:
                query = query.filter(Bill.salesman_id == salesman_id)
            
            bills = query.all()
            output.write('Bill Number,Type,Amount,Date,Vendor\n')
            for b in bills:
                output.write(f'{b.bill_number},{b.bill_type},{b.amount},{b.bill_date},{b.vendor.name if b.vendor else "N/A"}\n')
        
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=report_{report_type}_{datetime.now().strftime("%Y%m%d")}.csv'
        
        log_audit('EXPORT', 'Report', report_type, meta={'type': export_type, 'report': report_type})
        return response
    
    else:
        flash('PDF export coming soon', 'info')
        return redirect(request.referrer or url_for('analytics_dashboard'))

@app.route('/reports/delivery-efficiency')
@login_required
def delivery_efficiency_report():
    """Delivery efficiency report API endpoint"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = DeliveryOrder.query
    
    if date_from:
        query = query.filter(DeliveryOrder.delivery_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(DeliveryOrder.delivery_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    deliveries = query.all()
    
    total = len(deliveries)
    delivered = len([d for d in deliveries if d.status == 'delivered'])
    pending = len([d for d in deliveries if d.status == 'pending'])
    in_transit = len([d for d in deliveries if d.status == 'in_transit'])
    cancelled = len([d for d in deliveries if d.status == 'cancelled'])
    
    return jsonify({
        'total': total,
        'delivered': delivered,
        'pending': pending,
        'in_transit': in_transit,
        'cancelled': cancelled,
        'efficiency': (delivered / total * 100) if total > 0 else 0
    })

# OCR Verification Routes
@app.route('/ocr/verification')
@role_required('admin', 'computer_organiser')
def ocr_verification_dashboard():
    """OCR Verification Dashboard - Shows flagged bills with mismatches"""
    # Run verification check
    flagged_bills = run_daily_ocr_verification()
    
    # Also get bills that are pending verification
    pending_bills = Bill.query.filter(
        Bill.bill_type == 'normal',
        Bill.verification_status.in_(['pending', 'discrepancy'])
    ).all()
    
    return render_template('ocr_verification.html',
                         flagged_bills=flagged_bills,
                         pending_bills=pending_bills)

@app.route('/ocr/verify/<int:bill_id>', methods=['POST'])
@role_required('admin', 'computer_organiser')
def verify_ocr_bill_final(bill_id):
    """Final OCR verification - Approve, Correct, or Reject"""
    bill = Bill.query.get_or_404(bill_id)
    action = request.form.get('action')  # approve, correct, reject
    remarks = request.form.get('remarks', '')
    
    if action == 'approve':
        bill.verification_status = 'verified'
        status = 'Verified'
    elif action == 'correct':
        # Update bill with corrected data
        corrected_bill_number = request.form.get('corrected_bill_number', bill.bill_number)
        corrected_amount = request.form.get('corrected_amount', bill.amount)
        corrected_date = request.form.get('corrected_date')
        
        if corrected_bill_number != bill.bill_number:
            bill.bill_number = corrected_bill_number
        if corrected_amount:
            bill.amount = float(corrected_amount)
        if corrected_date:
            bill.bill_date = datetime.strptime(corrected_date, '%Y-%m-%d').date()
        
        bill.verification_status = 'verified'
        status = 'Corrected'
    elif action == 'reject':
        bill.verification_status = 'rejected'
        status = 'Rejected'
    else:
        flash('Invalid action', 'error')
        return redirect(url_for('ocr_verification_dashboard'))
    
    # Get mismatches for logging
    mismatches = detect_ocr_mismatches(bill)
    
    # Create verification log
    verification_log = OCRVerificationLog(
        bill_id=bill.id,
        verified_by=session['user_id'],
        status=status,
        mismatch_fields=json.dumps(mismatches) if mismatches else None,
        remarks=remarks
    )
    
    db.session.add(verification_log)
    db.session.commit()
    
    log_audit('VERIFY', 'OCR', bill.id, meta={
        'action': action,
        'status': status,
        'mismatches': mismatches,
        'remarks': remarks
    })
    
    flash(f'Bill {bill.bill_number} {status.lower()} successfully', 'success')
    return redirect(url_for('ocr_verification_dashboard'))

# Initialize database
def create_tables():
    db.create_all()
    
    # Create default users if none exist
    if not User.query.filter_by(role='admin').first():
        admin = User(
            username='admin', 
            email='admin@skanda.com', 
            role='admin',
            full_name='System Administrator',
            phone='+91-0000000000',
            is_active=True
        )
        admin.set_password('admin123')
        db.session.add(admin)
        print("✅ Default admin user created: username='admin', password='admin123'")
    
    # Create sample delivery man
    if not User.query.filter_by(role='delivery_man').first():
        delivery = User(
            username='delivery1',
            email='delivery@skanda.com',
            role='delivery_man',
            full_name='Delivery Man 1',
            phone='+91-1111111111',
            is_active=True
        )
        delivery.set_password('delivery123')
        db.session.add(delivery)
        print("✅ Default delivery man created: username='delivery1', password='delivery123'")
    
    # Create sample salesman
    if not User.query.filter_by(role='salesman').first():
        salesman = User(
            username='salesman1',
            email='salesman@skanda.com',
            role='salesman',
            full_name='Salesman 1',
            phone='+91-2222222222',
            is_active=True
        )
        salesman.set_password('salesman123')
        db.session.add(salesman)
        print("✅ Default salesman created: username='salesman1', password='salesman123'")
    
    # Create sample computer organiser
    if not User.query.filter_by(role='computer_organiser').first():
        organiser = User(
            username='organiser1',
            email='organiser@skanda.com',
            role='computer_organiser',
            full_name='Computer Organiser 1',
            phone='+91-3333333333',
            is_active=True
        )
        organiser.set_password('organiser123')
        db.session.add(organiser)
        print("✅ Default computer organiser created: username='organiser1', password='organiser123'")
    
        db.session.commit()

# Initialize database (lazy initialization for Vercel)
_db_initialized = False

def init_db():
    """Initialize database tables. Safe to call multiple times."""
    global _db_initialized
    if _db_initialized:
        return
    
    try:
        with app.app_context():
            # Warn about SQLite limitations on Vercel
            if os.environ.get('VERCEL') and app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite'):
                print("⚠️  WARNING: SQLite on Vercel uses /tmp which is ephemeral. Data will be lost between deployments.")
                print("⚠️  For production, use Vercel Postgres or another managed database service.")
            
            # Check if tables exist by trying to query
            try:
                User.query.limit(1).all()
                _db_initialized = True
                return
            except Exception:
                pass
            
            # Create tables if they don't exist
            create_tables()
            _db_initialized = True
    except Exception as e:
        print(f"Database initialization error: {e}")
        import traceback
        traceback.print_exc()
        # Don't set _db_initialized to True on error, allow retry

# Initialize database on first request (Flask 2.3+ compatible)
@app.before_request
def initialize_database():
    """Initialize database before first request (lazy initialization)."""
    if not _db_initialized:
        init_db()

# Ensure static files are served with proper headers (for Vercel compatibility)
@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files with proper cache headers."""
    from flask import send_from_directory
    response = send_from_directory(app.static_folder, filename)
    # Add cache headers for static assets
    if filename.endswith(('.css', '.js', '.png', '.jpg', '.jpeg', '.gif', '.svg', '.ico', '.woff', '.woff2', '.ttf', '.eot')):
        response.cache_control.max_age = 31536000  # 1 year
        response.cache_control.public = True
    return response

# For local development
if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
else:
    # For Vercel/serverless: Database will be initialized on first request via @app.before_request
    pass
